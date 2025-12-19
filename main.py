import cv2
import mediapipe as mp
import os
from datetime import datetime
import calendar
from openpyxl import Workbook, load_workbook
import streamlit as st
from streamlit_webrtc import webrtc_streamer, VideoProcessorBase

# ---------------- PAGE SETUP ----------------
st.set_page_config(page_title="Face Attendance System", layout="centered")
st.title("😊 Good Morning! Have a Nice Day")
st.write("Enter your name and allow camera access")

# ---------------- SESSION STATE ----------------
if "marked" not in st.session_state:
    st.session_state["marked"] = False

# ---------------- EXCEL FILE ----------------
excel_file = "Attendance.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Day", "Time"])
    wb.save(excel_file)

# ---------------- FACE DETECTION ----------------
def get_face_detector():
    mp_face = mp.solutions.face_detection
    return mp_face.FaceDetection(
        model_selection=0,
        min_detection_confidence=0.5
    )

face_detection = get_face_detector()

# ---------------- ATTENDANCE FUNCTION ----------------
def mark_attendance(name):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    day = calendar.day_name[now.weekday()]
    time = now.strftime("%H:%M:%S")

    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == date:
            return False

    ws.append([name, date, day, time])
    wb.save(excel_file)
    return True

# ---------------- WEBRTC PROCESSOR ----------------
class FaceProcessor(VideoProcessorBase):
    def recv(self, frame):
        img = frame.to_ndarray(format="bgr24")
        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        results = face_detection.process(rgb)

        if results.detections and not st.session_state["marked"]:
            saved = mark_attendance(st.session_state["name"])
            st.session_state["marked"] = True

            if saved:
                st.success(f"✔ Attendance marked for {st.session_state['name']}")
            else:
                st.warning(f"⚠ {st.session_state['name']} already marked today")

        return img

# ---------------- STREAMLIT UI ----------------
name = st.text_input("Enter Your Name", key="name")

start = st.checkbox("Start Attendance")

if start:
    if name.strip() == "":
        st.warning("Please enter your name")
    else:
        webrtc_streamer(
            key="face-attendance",
            video_processor_factory=FaceProcessor,
            media_stream_constraints={"video": True, "audio": False},
        )
